from django.test import TestCase
from django.contrib.auth import get_user_model
from django.urls import reverse

from io import BytesIO

from openpyxl import load_workbook

from .models import Tblcourse, CourseSchedule, Quiz, QuizQuestion
from apps.students.models import Tblstudents


class CourseScheduleTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(username='tester', password='secret123')
        self.course = Tblcourse.objects.create(name='BSIT', section='1A', schoolyr='2024-2025', user=self.user)

    def test_schedule_save_creates_schedule(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('courses:schedule_save_ajax'),
            {
                'course': self.course.courseid,
                'day': '1',
                'start_time': '08:00',
                'end_time': '09:30',
                'room': 'Room 201',
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'success')
        schedule = CourseSchedule.objects.latest('id')
        self.assertEqual(schedule.course, self.course)
        self.assertEqual(schedule.day, 1)
        self.assertEqual(schedule.room, 'Room 201')

    def test_schedule_save_requires_course_and_valid_times(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('courses:schedule_save_ajax'),
            {'course': '', 'day': '1', 'start_time': '08:00', 'end_time': '09:30'},
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('Course', response.json()['error'])

        response = self.client.post(
            reverse('courses:schedule_save_ajax'),
            {
                'course': self.course.courseid,
                'day': '1',
                'start_time': '10:00',
                'end_time': '09:30',
            },
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('End time', response.json()['error'])

    def test_schedule_events_ajax_maps_days_and_formats_times(self):
        self.client.force_login(self.user)
        schedule = CourseSchedule.objects.create(
            course=self.course, day=1, start_time='08:00', end_time='09:30', room='201'
        )
        response = self.client.get(reverse('courses:schedule_events_ajax'))
        self.assertEqual(response.status_code, 200)
        events = response.json()
        self.assertEqual(len(events), 1)
        self.assertEqual(events[0]['id'], schedule.id)
        self.assertEqual(events[0]['daysOfWeek'], [1])
        self.assertEqual(events[0]['startTime'], '08:00:00')
        self.assertEqual(events[0]['endTime'], '09:30:00')
        self.assertEqual(events[0]['extendedProps']['room'], '201')

    def test_schedule_delete_removes_schedule(self):
        self.client.force_login(self.user)
        schedule = CourseSchedule.objects.create(
            course=self.course, day=2, start_time='10:00', end_time='11:00'
        )
        response = self.client.post(reverse('courses:schedule_delete_ajax', args=[schedule.id]))
        self.assertEqual(response.json()['status'], 'deleted')
        self.assertFalse(CourseSchedule.objects.filter(pk=schedule.id).exists())

    def test_schedule_endpoints_require_login(self):
        response = self.client.get(reverse('courses:schedule_events_ajax'))
        self.assertEqual(response.status_code, 302)


class QuizQuestionSaveTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(username='tester', password='secret123')
        self.course = Tblcourse.objects.create(name='Test Course', section='A', schoolyr='2024-2025', user=self.user)
        self.quiz = Quiz.objects.create(course=self.course, title='Sample Quiz')

    def test_identification_question_saves_text_answer(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('courses:question_save', args=[self.quiz.id]),
            {
                'question_text': 'What is the capital of France?',
                'question_type': 'identification',
                'identification_answer': 'Paris',
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        question = QuizQuestion.objects.latest('id')
        self.assertEqual(question.question_type, 'identification')
        self.assertEqual(question.correct_option, 'Paris')
        self.assertEqual(question.option_a, '')

    def test_true_false_question_uses_boolean_style_answer(self):
        self.client.force_login(self.user)
        self.client.post(
            reverse('courses:question_save', args=[self.quiz.id]),
            {
                'question_text': 'The Earth is round.',
                'question_type': 'true_false',
                'correct_option': 'False',
            },
            follow=True,
        )
        question = QuizQuestion.objects.latest('id')
        self.assertEqual(question.question_type, 'true_false')
        self.assertEqual(question.correct_option, 'False')
        self.assertEqual(question.option_a, 'True')
        self.assertEqual(question.option_b, 'False')


class CourseStudentsExportTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(username='tester', password='secret123')
        self.course = Tblcourse.objects.create(name='BSIT', section='1A', schoolyr='2024-2025', user=self.user)
        Tblstudents.objects.create(idno='2024-001', fullname='Alice Doe', courseid=str(self.course.courseid), user=self.user)
        Tblstudents.objects.create(idno='2024-002', fullname='Bob Smith', courseid=str(self.course.courseid), user=self.user)
        Tblstudents.objects.create(idno='2024-003', fullname='Carol Lee', courseid='999', user=self.user)

    def test_export_returns_excel_with_students(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('courses:course_students_export', args=[self.course.courseid]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertEqual(sheet['A1'].value, 'BSIT - 1A')
        self.assertEqual(sheet['A2'].value, 'School Year: 2024-2025')
        self.assertEqual(sheet.cell(row=3, column=3).value, 'Student Name')
        self.assertEqual(sheet['A4'].value, 1)
        self.assertEqual(sheet['B4'].value, '2024-001')
        self.assertEqual(sheet['C4'].value, 'Alice Doe')
        self.assertEqual(sheet['A5'].value, 2)
        self.assertEqual(sheet['C5'].value, 'Bob Smith')
        self.assertIsNone(sheet['A6'].value)

    def test_export_requires_login(self):
        response = self.client.get(reverse('courses:course_students_export', args=[self.course.courseid]))
        self.assertEqual(response.status_code, 302)
