import json
import re
from datetime import datetime

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db import IntegrityError
from django.db.models import Q, Count
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Tblcourse, CourseSchedule, Quiz, QuizQuestion
from apps.students.models import Tblstudents
from apps.scoping import (
    scoped_course,
    scoped_courses,
    scoped_quizzes,
    scoped_quiz_questions,
    scoped_schedules,
    scoped_students,
)

@login_required(login_url='login')
def course_list_page(request):
    query = request.GET.get('q', '').strip()

    courses_list = scoped_courses(request.user).order_by('name')

    if query:
        courses_list = courses_list.filter(
            Q(name__icontains=query) |
            Q(courseid__icontains=query) |
            Q(section__icontains=query) |
            Q(schoolyr__icontains=query)
        )

    paginator = Paginator(courses_list, 10)
    page_number = request.GET.get('page', 1)
    courses = paginator.get_page(page_number)

    # courseid on Tblcourse is int; courseid on Tblstudents is varchar,
    # so cast to str for a clean match against the student table.
    course_ids = [str(c.courseid) for c in courses.object_list]
    counts = (
        scoped_students(request.user)
        .filter(courseid__in=course_ids)
        .values('courseid')
        .annotate(student_count=Count('id'))
    )
    counts_map = {row['courseid']: row['student_count'] for row in counts}

    for course in courses.object_list:
        course.student_count = counts_map.get(str(course.courseid), 0)

    return render(request, 'courses.html', {
        'courses': courses,
        'query': query,
    })


@login_required(login_url='login')
def course_list_ajax(request):
    courses_list = scoped_courses(request.user).order_by('name')
    paginator = Paginator(courses_list, 10)
    page_number = request.GET.get('page', 1)
    courses = paginator.get_page(page_number)
    data = list(courses.object_list.values('courseid', 'name', 'section', 'schoolyr'))
    return JsonResponse({
        'data': data,
        'has_next': courses.has_next(),
        'has_prev': courses.has_previous(),
        'page': courses.number,
        'pages': courses.paginator.num_pages,
        'total': courses.paginator.count,
    })


@login_required(login_url='login')
def course_get_ajax(request, pk):
    course = scoped_course(request.user, pk)
    return JsonResponse({
        'courseid': course.courseid,
        'name': course.name,
        'section': course.section,
        'schoolyr': course.schoolyr,
    })


@login_required(login_url='login')
@require_POST
def course_save_ajax(request):
    course_id = request.POST.get('courseid', '').strip()

    if course_id:
        course = scoped_course(request.user, course_id)
    else:
        course = Tblcourse(user=request.user)

    name = request.POST.get('name', '').strip()
    section = request.POST.get('section', '').strip()
    schoolyr = request.POST.get('schoolyr', '').strip()
    status = request.POST.get('status', 'active').strip().lower()

    if not name:
        return JsonResponse({'error': 'Course name is required.'}, status=400)
    if not section:
        return JsonResponse({'error': 'Section is required.'}, status=400)
    if not schoolyr:
        return JsonResponse({'error': 'School year is required.'}, status=400)
    if status not in dict(Tblcourse.STATUS_CHOICES):
        status = 'active'

    course.name = name
    course.section = section
    course.schoolyr = schoolyr
    course.status = status

    try:
        course.save()
    except IntegrityError:
        return JsonResponse({'error': 'Could not save course due to a conflict.'}, status=400)

    return JsonResponse({
        'status': 'success',
        'courseid': course.courseid,
        'name': course.name,
        'section': course.section,
        'schoolyr': course.schoolyr,
        'status': course.status,
    })


@login_required(login_url='login')
def course_students_export(request, pk):
    course = scoped_course(request.user, pk)
    students = scoped_students(request.user).filter(courseid=str(course.courseid)).order_by('fullname')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Students'

    ws.merge_cells('A1:C1')
    ws['A1'] = f'{course.name} - {course.section}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    ws['A2'] = f'School Year: {course.schoolyr}'
    ws['A2'].font = Font(italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['#', 'ID Number', 'Student Name']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for index, student in enumerate(students, start=1):
        row = index + 3
        ws.cell(row=row, column=1, value=index)
        ws.cell(row=row, column=2, value=student.idno)
        ws.cell(row=row, column=3, value=student.fullname)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.freeze_panes = 'A4'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = re.sub(r'[^\w\s-]', '', course.name).strip() or 'course'
    filename = f'{safe_name}_students.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url='login')
@require_POST
def course_delete_ajax(request, pk):
    course = scoped_course(request.user, pk)
    course.delete()
    return JsonResponse({'status': 'deleted'})


@login_required(login_url='login')
def schedule_page(request):
    courses = scoped_courses(request.user).order_by('name')
    selected_course_id = request.GET.get('course', '').strip()
    schedules_qs = scoped_schedules(request.user)
    if selected_course_id:
        schedules_qs = schedules_qs.filter(course_id=selected_course_id)
    schedules = schedules_qs.order_by('course__name', 'day', 'start_time')

    return render(request, 'courses/schedules.html', {
        'courses': courses,
        'schedules': schedules,
        'selected_course_id': selected_course_id,
        'schedule_day_choices': CourseSchedule.DAY_CHOICES,
    })


@login_required(login_url='login')
def schedule_get_ajax(request, pk):
    schedule = get_object_or_404(scoped_schedules(request.user), pk=pk)
    return JsonResponse({
        'id': schedule.id,
        'course': schedule.course_id,
        'day': schedule.day,
        'start_time': schedule.start_time.strftime('%H:%M'),
        'end_time': schedule.end_time.strftime('%H:%M'),
        'room': schedule.room,
    })


def _parse_time(value, label):
    try:
        return datetime.strptime(value, '%H:%M').time()
    except (TypeError, ValueError):
        return None


@login_required(login_url='login')
@require_POST
def schedule_save_ajax(request):
    schedule_id = request.POST.get('id', '').strip()
    course_id = request.POST.get('course', '').strip()
    day_raw = request.POST.get('day', '').strip()
    start_raw = request.POST.get('start_time', '').strip()
    end_raw = request.POST.get('end_time', '').strip()
    room = request.POST.get('room', '').strip()

    if not course_id:
        return JsonResponse({'error': 'Course is required.'}, status=400)
    course = scoped_course(request.user, course_id)

    try:
        day = int(day_raw)
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Day is required.'}, status=400)
    if day not in dict(CourseSchedule.DAY_CHOICES):
        return JsonResponse({'error': 'Invalid day.'}, status=400)

    start_time = _parse_time(start_raw, 'Start time')
    end_time = _parse_time(end_raw, 'End time')
    if start_time is None:
        return JsonResponse({'error': 'Start time is required.'}, status=400)
    if end_time is None:
        return JsonResponse({'error': 'End time is required.'}, status=400)
    if end_time <= start_time:
        return JsonResponse({'error': 'End time must be after the start time.'}, status=400)

    if schedule_id:
        schedule = get_object_or_404(scoped_schedules(request.user), pk=schedule_id)
    else:
        schedule = CourseSchedule()

    schedule.course = course
    schedule.day = day
    schedule.start_time = start_time
    schedule.end_time = end_time
    schedule.room = room

    try:
        schedule.save()
    except IntegrityError:
        return JsonResponse({'error': 'Could not save schedule due to a conflict.'}, status=400)

    return JsonResponse({
        'status': 'success',
        'id': schedule.id,
        'course': schedule.course_id,
        'course_name': schedule.course.name,
        'section': schedule.course.section,
        'day': schedule.day,
        'day_label': schedule.get_day_display(),
        'start_time': schedule.start_time.strftime('%H:%M'),
        'end_time': schedule.end_time.strftime('%H:%M'),
        'room': schedule.room,
    })


@login_required(login_url='login')
@require_POST
def schedule_delete_ajax(request, pk):
    schedule = get_object_or_404(scoped_schedules(request.user), pk=pk)
    schedule.delete()
    return JsonResponse({'status': 'deleted'})


@login_required(login_url='login')
def schedule_events_ajax(request):
    schedules = scoped_schedules(request.user).filter(course__status='active')

    events = []
    for schedule in schedules:
        events.append({
            'id': schedule.id,
            'title': f"{schedule.course.name} - {schedule.course.section}",
            'daysOfWeek': [schedule.day % 7],
            'startTime': schedule.start_time.strftime('%H:%M:%S'),
            'endTime': schedule.end_time.strftime('%H:%M:%S'),
            'backgroundColor': '#d8f3dc',
            'borderColor': '#a3d9b1',
            'textColor': '#1b4332',
            'extendedProps': {
                'room': schedule.room,
                'course': schedule.course.name,
                'section': schedule.course.section,
            },
        })

    return JsonResponse(events, safe=False)


@login_required(login_url='login')
def quiz_list_page(request, course_id):
    course = scoped_course(request.user, course_id)
    quizzes = course.quizzes.all().order_by('-created_at')
    selected_quiz_id = request.GET.get('quiz_id', '').strip()
    selected_quiz = None

    if selected_quiz_id:
        selected_quiz = course.quizzes.filter(pk=selected_quiz_id).first()

    return render(request, 'courses/quizzes.html', {
        'course': course,
        'quizzes': quizzes,
        'selected_quiz': selected_quiz,
    })


@login_required(login_url='login')
def quiz_preview_page(request, quiz_id):
    quiz = get_object_or_404(scoped_quizzes(request.user), pk=quiz_id)
    questions = quiz.questions.all().order_by('order', 'id')
    questions_data = [
        {
            'id': question.id,
            'text': question.question_text,
            'type': question.question_type,
            'option_a': question.option_a,
            'option_b': question.option_b,
            'option_c': question.option_c,
            'option_d': question.option_d,
            'correct': question.correct_option,
        }
        for question in questions
    ]

    return render(request, 'courses/quiz_preview.html', {
        'quiz': quiz,
        'questions': questions,
        'questions_data': json.dumps(questions_data),
    })


@login_required(login_url='login')
@require_POST
def quiz_save_view(request, course_id):
    course = scoped_course(request.user, course_id)
    quiz_id = request.POST.get('quiz_id', '').strip()

    if quiz_id:
        quiz = get_object_or_404(scoped_quizzes(request.user), pk=quiz_id, course=course)
    else:
        quiz = Quiz(course=course)

    title = request.POST.get('title', '').strip()
    description = request.POST.get('description', '').strip()
    is_active = request.POST.get('is_active') == 'on'

    if not title:
        return redirect('courses:quiz_list', course_id=course.courseid)

    quiz.title = title
    quiz.description = description
    quiz.is_active = is_active
    quiz.save()

    return redirect('courses:quiz_list', course_id=course.courseid)


@login_required(login_url='login')
@require_POST
def quiz_delete_view(request, course_id, quiz_id):
    quiz = get_object_or_404(scoped_quizzes(request.user), pk=quiz_id, course__courseid=course_id)
    quiz.delete()
    return redirect('courses:quiz_list', course_id=course_id)


@login_required(login_url='login')
@require_POST
def question_save_view(request, quiz_id):
    quiz = get_object_or_404(scoped_quizzes(request.user), pk=quiz_id)
    question_text = request.POST.get('question_text', '').strip()
    question_type = request.POST.get('question_type', 'multiple_choice').strip().lower()
    option_a = request.POST.get('option_a', '').strip()
    option_b = request.POST.get('option_b', '').strip()
    option_c = request.POST.get('option_c', '').strip()
    option_d = request.POST.get('option_d', '').strip()
    correct_option = request.POST.get('correct_option', '').strip()
    identification_answer = request.POST.get('identification_answer', '').strip()

    if question_type == 'true_false':
        option_a = 'True'
        option_b = 'False'
        if correct_option not in {'True', 'False'}:
            correct_option = 'True'
    elif question_type == 'identification':
        correct_option = identification_answer
        option_a = ''
        option_b = ''
        option_c = ''
        option_d = ''

    if question_type == 'identification':
        if not question_text or not correct_option:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax') == '1':
                return JsonResponse({'error': 'Question text and answer are required.'}, status=400)
            return redirect('courses:quiz_list', course_id=quiz.course.courseid)
    else:
        if not question_text or not option_a or not option_b:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax') == '1':
                return JsonResponse({'error': 'Question text and options are required.'}, status=400)
            return redirect('courses:quiz_list', course_id=quiz.course.courseid)

    QuizQuestion.objects.create(
        quiz=quiz,
        question_text=question_text,
        question_type=question_type,
        option_a=option_a,
        option_b=option_b,
        option_c=option_c,
        option_d=option_d,
        correct_option=correct_option,
        order=quiz.questions.count() + 1,
    )

    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax') == '1':
        questions_html = render_to_string(
            'courses/partials/question_list.html',
            {'questions': quiz.questions.all().order_by('order', 'id')},
            request=request,
        )
        return JsonResponse({
            'status': 'success',
            'question_count': quiz.questions.count(),
            'questions_html': questions_html,
        })

    return redirect('courses:quiz_list', course_id=quiz.course.courseid)


@login_required(login_url='login')
@require_POST
def question_delete_view(request, question_id):
    question = get_object_or_404(scoped_quiz_questions(request.user), pk=question_id)
    course_id = question.quiz.course.courseid
    question.delete()
    return redirect('courses:quiz_list', course_id=course_id)
